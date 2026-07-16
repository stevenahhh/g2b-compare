"""Import only manifest-pinned workbook relationships."""

from __future__ import annotations

import hashlib
import json
from contextlib import closing
from dataclasses import dataclass
from enum import StrEnum
from typing import TYPE_CHECKING, Final, Literal, override

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.worksheet.worksheet import Worksheet

from . import RelationSourceManifest

PINNED_FILENAME: Final = (
    "250725-전남 광양시 아트케이션 관광스테이 확충사업 CCTV 설비 내역서.xlsx"
)
PINNED_SOURCE_SHA256: Final = (
    "445012e259ab5318a1d52468cce93ee28a55a8bcb467876f40a47a939e4668db"
)
_MANIFEST: Final = "7088aa516249bd06ee98db8be75e78b70d3e87605bd40ca8b7c4c8690ece4f79"


PINNED_RELATION_MANIFEST: Final = RelationSourceManifest(
    source_sha256=PINNED_SOURCE_SHA256,
    sheet_name="자재내역서",
    headers=(
        ("B8", "1-1. 본품"),
        ("B10", "1-2. 우수제품 옵션품목"),
        ("B26", "2-1. 옵션품목"),
    ),
    parent_coordinate="N9",
    parent_value="24684676",
    child_coordinates=tuple(f"N{row}" for row in range(11, 23)),
    unbound_coordinates=tuple(f"N{row}" for row in range(27, 30)),
    curated_relationship_count=12,
    unbound_option_count=3,
)


class RelationImportFailure(StrEnum):
    """Exhaustive fail-closed workbook relation outcomes."""

    SOURCE_HASH_MISMATCH = "source_hash_mismatch"
    MANIFEST_DRIFT = "manifest_drift"
    SHEET_MISSING = "sheet_missing"
    HEADER_DRIFT = "header_drift"
    FORMULA_ID = "formula_id"
    MERGED_ID = "merged_id"
    MISSING_PARENT = "missing_parent"
    MISSING_CHILD = "missing_child"
    BAD_ID = "bad_id"
    SELF_LINK = "self_link"
    DUPLICATE_ROW = "duplicate_row"
    ZERO_IMPORT = "zero_import"
    UNEXPECTED_RELATION_COUNT = "unexpected_relation_count"


@dataclass(frozen=True, slots=True)
class RelationImportError(Exception):
    """Reject workbook relation input at its first invalid boundary."""

    failure: RelationImportFailure
    coordinate: str | None = None

    @override
    def __str__(self) -> str:
        return f"workbook relation {self.failure.value}:{self.coordinate or ''}"


@dataclass(frozen=True, slots=True)
class CuratedRelation:
    """One explicit parent-child relation with workbook provenance."""

    relation_id: str
    parent_id: str
    child_id: str
    source_sha: str
    sheet_name: str
    row_no: int
    source_type: Literal["curated_workbook"] = "curated_workbook"


@dataclass(frozen=True, slots=True)
class UnboundOption:
    """One option cell excluded because no explicit parent binds it."""

    product_id: str
    source_sha: str
    sheet_name: str
    row_no: int
    reason: Literal["unbound_option"] = "unbound_option"


@dataclass(frozen=True, slots=True)
class RelationSnapshotCandidate:
    """Immutable candidate that Todo 12 may select for a release."""

    source_manifest_sha: str
    relation_content_sha: str
    status: Literal["complete"] = "complete"


@dataclass(frozen=True, slots=True)
class WorkbookRelationImport:
    """Complete relation import or an explicit non-target workbook result."""

    snapshot: RelationSnapshotCandidate | None
    relations: tuple[CuratedRelation, ...]
    quarantined: tuple[UnboundOption, ...]


def import_workbook_relations(
    path: Path,
    manifest: RelationSourceManifest = PINNED_RELATION_MANIFEST,
) -> WorkbookRelationImport:
    """Import one trusted workbook without mutating the source file."""
    if manifest != PINNED_RELATION_MANIFEST:
        raise RelationImportError(RelationImportFailure.MANIFEST_DRIFT)
    return _import_verified_workbook(path, manifest)


def _import_verified_workbook(
    path: Path,
    manifest: RelationSourceManifest,
) -> WorkbookRelationImport:
    source_sha = hashlib.sha256(path.read_bytes()).hexdigest()
    if source_sha != manifest.source_sha256:
        if path.name == PINNED_FILENAME:
            raise RelationImportError(RelationImportFailure.SOURCE_HASH_MISMATCH)
        return WorkbookRelationImport(None, (), ())
    if (
        len(manifest.child_coordinates) != manifest.curated_relationship_count
        or len(manifest.unbound_coordinates) != manifest.unbound_option_count
    ):
        raise RelationImportError(RelationImportFailure.UNEXPECTED_RELATION_COUNT)
    with closing(
        load_workbook(path, read_only=False, data_only=False, keep_links=True)
    ) as workbook:
        if manifest.sheet_name not in workbook.sheetnames:
            raise RelationImportError(RelationImportFailure.SHEET_MISSING)
        worksheet = workbook[manifest.sheet_name]
        _validate_headers(worksheet, manifest.headers)
        if all(
            not _cell_text(worksheet, coordinate)
            for coordinate in manifest.child_coordinates
        ):
            raise RelationImportError(RelationImportFailure.ZERO_IMPORT)
        parent_id = _product_id(
            worksheet,
            manifest.parent_coordinate,
            RelationImportFailure.MISSING_PARENT,
        )
        if parent_id != manifest.parent_value:
            raise RelationImportError(
                RelationImportFailure.HEADER_DRIFT,
                manifest.parent_coordinate,
            )
        child_ids = _child_ids(worksheet, parent_id, manifest.child_coordinates)
        relations = tuple(
            CuratedRelation(
                relation_id=_relation_id(
                    parent_id,
                    child_id,
                    (source_sha, manifest.sheet_name, int(coordinate[1:])),
                ),
                parent_id=parent_id,
                child_id=child_id,
                source_sha=source_sha,
                sheet_name=manifest.sheet_name,
                row_no=int(coordinate[1:]),
            )
            for coordinate, child_id in zip(
                manifest.child_coordinates,
                child_ids,
                strict=True,
            )
        )
        quarantined = tuple(
            UnboundOption(
                product_id=_product_id(
                    worksheet,
                    coordinate,
                    RelationImportFailure.UNEXPECTED_RELATION_COUNT,
                ),
                source_sha=source_sha,
                sheet_name=manifest.sheet_name,
                row_no=int(coordinate[1:]),
            )
            for coordinate in manifest.unbound_coordinates
        )
        unbound_ids = tuple(item.product_id for item in quarantined)
        if len(set(unbound_ids)) != len(unbound_ids) or set(child_ids).intersection(
            unbound_ids
        ):
            raise RelationImportError(RelationImportFailure.DUPLICATE_ROW)
    content = tuple(
        (
            relation.relation_id,
            relation.parent_id,
            relation.child_id,
            relation.source_type,
            relation.source_sha,
            relation.sheet_name,
            relation.row_no,
        )
        for relation in relations
    )
    content_json = json.dumps(content, ensure_ascii=False, separators=(",", ":"))
    content_sha = hashlib.sha256(content_json.encode()).hexdigest()
    snapshot = RelationSnapshotCandidate(_MANIFEST, content_sha)
    return WorkbookRelationImport(snapshot, relations, quarantined)


def _validate_headers(
    worksheet: Worksheet,
    headers: tuple[tuple[str, str], ...],
) -> None:
    for coordinate, expected in headers:
        if _cell_text(worksheet, coordinate) != expected:
            raise RelationImportError(RelationImportFailure.HEADER_DRIFT, coordinate)


def _cell_text(worksheet: Worksheet, coordinate: str) -> str:
    value = worksheet.cell(*coordinate_to_tuple(coordinate)).value
    return "" if value is None else str(value).strip()


def _product_id(
    worksheet: Worksheet,
    coordinate: str,
    blank_failure: RelationImportFailure,
) -> str:
    if coordinate in worksheet.merged_cells:
        raise RelationImportError(RelationImportFailure.MERGED_ID, coordinate)
    data_type = worksheet.cell(*coordinate_to_tuple(coordinate)).data_type
    if data_type == "f":
        raise RelationImportError(RelationImportFailure.FORMULA_ID, coordinate)
    value = _cell_text(worksheet, coordinate)
    if not value:
        raise RelationImportError(blank_failure, coordinate)
    if (
        len(value) != len(PINNED_RELATION_MANIFEST.parent_value)
        or not value.isascii()
        or not value.isdigit()
    ):
        raise RelationImportError(RelationImportFailure.BAD_ID, coordinate)
    return value


def _child_ids(
    worksheet: Worksheet,
    parent_id: str,
    coordinates: tuple[str, ...],
) -> tuple[str, ...]:
    child_ids: list[str] = []
    for coordinate in coordinates:
        child_id = _product_id(
            worksheet,
            coordinate,
            RelationImportFailure.MISSING_CHILD,
        )
        if child_id == parent_id:
            raise RelationImportError(RelationImportFailure.SELF_LINK, coordinate)
        if child_id in child_ids:
            raise RelationImportError(RelationImportFailure.DUPLICATE_ROW, coordinate)
        child_ids.append(child_id)
    return tuple(child_ids)


def _relation_id(
    parent_id: str,
    child_id: str,
    source_identity: tuple[str, str, int],
) -> str:
    identity = json.dumps(
        (parent_id, child_id, *source_identity),
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode()
    return hashlib.sha256(identity).hexdigest()
