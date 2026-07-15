"""Fail closed when workbook inputs contain unsafe or changed source data."""

from __future__ import annotations

import hashlib
from contextlib import closing
from dataclasses import dataclass
from typing import TYPE_CHECKING, Final, override
from zipfile import ZipFile

if TYPE_CHECKING:
    from pathlib import Path

_EXTERNAL_LINK_DIRECTORY: Final = "xl/externallinks/"
_EXTERNAL_LINK_RELATION: Final = b"/relationships/externallink"


@dataclass(frozen=True, slots=True)
class WorkbookFormulaError(Exception):
    """Report the first formula rejected by workbook inspection."""

    path: Path
    coordinate: str

    @override
    def __str__(self) -> str:
        return f"workbook formula error: {self.path.name}:{self.coordinate}"


@dataclass(frozen=True, slots=True)
class WorkbookExternalLinkError(Exception):
    """Report the first external reference rejected by workbook inspection."""

    path: Path
    coordinate: str

    @override
    def __str__(self) -> str:
        return f"workbook external link: {self.path.name}:{self.coordinate}"


@dataclass(frozen=True, slots=True)
class WorkbookSourceHashError(Exception):
    """Report a source workbook whose SHA-256 changed."""

    path: Path
    expected: str
    actual: str

    @override
    def __str__(self) -> str:
        return f"source workbook SHA changed: {self.path.name}"


def inspect_workbook(path: Path) -> None:
    """Reject formulas and external references before extraction."""
    with ZipFile(path) as package:
        for member in package.namelist():
            normalized_member = member.replace("\\", "/").casefold().lstrip("/")
            if normalized_member.startswith(_EXTERNAL_LINK_DIRECTORY):
                raise WorkbookExternalLinkError(path, member)
            if normalized_member.endswith(".rels") and (
                _EXTERNAL_LINK_RELATION in package.read(member).lower()
            ):
                raise WorkbookExternalLinkError(path, member)

    from openpyxl import load_workbook  # noqa: PLC0415

    with closing(
        load_workbook(path, read_only=True, data_only=False, keep_links=True)
    ) as workbook:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type != "f":
                        continue
                    formula = str(cell.value)
                    if "[" in formula and "]" in formula:
                        raise WorkbookExternalLinkError(path, cell.coordinate)
                    raise WorkbookFormulaError(path, cell.coordinate)


def verify_source_hash(path: Path, expected: str) -> None:
    """Reject source bytes that differ from a pinned SHA-256."""
    actual = hashlib.sha256(path.read_bytes()).hexdigest()
    if actual != expected:
        raise WorkbookSourceHashError(path, expected, actual)
