"""Read priority companies and options without modifying their workbook."""

from __future__ import annotations

from datetime import date, datetime
from typing import TYPE_CHECKING, Final, Literal, cast

from openpyxl import load_workbook

from g2b_compare.priority_models import (
    PriorityCompany,
    PriorityDataset,
    PriorityOption,
)

if TYPE_CHECKING:
    from collections.abc import Iterable
    from pathlib import Path

COMPANY_HEADERS: Final = (
    "No",
    "계약업체",
    "본사소재지",
    "기업구분",
    "상품수",
    "계약종료일",
)
OPTION_HEADERS: Final = (
    "업체명",
    "구분",
    "조달식별번호",
    "품명",
    "규격",
    "금액",
    "기타상세",
)
type CellValue = str | int | float | date | datetime | None


def read_priority_workbook(path: Path) -> PriorityDataset:
    """Parse the two source sheets and retain every populated option row."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        companies_sheet = workbook["업체소재별현황"]
        options_sheet = workbook["우수옵션"]
        _require_headers(
            companies_sheet.iter_rows(min_row=5, max_row=5, values_only=True),
            COMPANY_HEADERS,
        )
        _require_headers(
            options_sheet.iter_rows(min_row=2, max_row=2, values_only=True),
            OPTION_HEADERS,
        )

        companies: list[PriorityCompany] = []
        for row_number, raw in enumerate(
            companies_sheet.iter_rows(min_row=6, values_only=True), start=6
        ):
            values = _values(raw, len(COMPANY_HEADERS))
            if not _text(values[1]):
                continue
            companies.append(
                PriorityCompany(
                    source_row=row_number,
                    name=_text(values[1]),
                    location=_text(values[2]),
                    company_type=_text(values[3]),
                    declared_product_count=_integer(values[4]),
                    contract_end_date=_date_text(values[5]),
                )
            )

        options: list[PriorityOption] = []
        for row_number, raw in enumerate(
            options_sheet.iter_rows(min_row=3, values_only=True), start=3
        ):
            values = _values(raw, len(OPTION_HEADERS))
            if not _text(values[0]):
                continue
            options.append(
                PriorityOption(
                    source_row=row_number,
                    company_name=_text(values[0]),
                    kind=_kind(values[1]),
                    product_id=_product_id(values[2]),
                    item_name=_text(values[3]),
                    spec=_text(values[4]),
                    price_won=_integer(values[5]),
                    details=_text(values[6]),
                )
            )
    finally:
        workbook.close()
    return PriorityDataset(companies=tuple(companies), options=tuple(options))


def _require_headers(
    rows: Iterable[tuple[object, ...]], expected: tuple[str, ...]
) -> None:
    raw = next(iter(rows))
    actual = tuple(_text(value) for value in raw)
    if actual[: len(expected)] != expected:
        msg = f"unexpected workbook headers: {actual[: len(expected)]!r}"
        raise ValueError(msg)


def _values(raw: tuple[object, ...], width: int) -> tuple[CellValue, ...]:
    values = tuple(cast("CellValue", value) for value in raw[:width])
    return (*values, *((None,) * (width - len(values))))


def _text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _integer(value: object) -> int:
    text = _text(value).replace(",", "")
    return int(float(text)) if text else 0


def _product_id(value: object) -> str:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return f"{int(value):08d}"
    return _text(value)


def _kind(value: object) -> Literal["추가선택", "선택부품"]:
    match _text(value):
        case "추가선택":
            return "추가선택"
        case "선택부품":
            return "선택부품"
        case invalid:
            msg = f"unexpected option kind: {invalid!r}"
            raise ValueError(msg)


def _date_text(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _text(value)
