"""Fill the fixed estimate workbook without disturbing legacy drawings."""

from __future__ import annotations

import hashlib
from pathlib import Path
from typing import TYPE_CHECKING, Final, Literal, final
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import load_workbook

from g2b_compare.db.connection import connect
from g2b_compare.db.sql import as_int, as_text, query

from .estimate_export_models import (
    ComparisonSnapshot,
    EstimateExportError,
    TemplateManifest,
)
from .estimate_store import EstimateStore
from .estimate_xlsx import write_draft

if TYPE_CHECKING:
    from .estimate_models import EstimateDraft

ASSET_DIRECTORY: Final = Path(__file__).resolve().parents[1] / "assets"
DEFAULT_TEMPLATE: Final = ASSET_DIRECTORY / "estimate-template-v1.xlsx"
DEFAULT_MANIFEST: Final = ASSET_DIRECTORY / "estimate-template-v1.json"
NO_IMAGE: Final = ASSET_DIRECTORY / "estimate-no-image.png"
TEMPLATE_OVERWRITE: Final = "기준 템플릿은 덮어쓸 수 없음"
TEMPLATE_HASH_CHANGED: Final = "기준 템플릿 해시가 일치하지 않음"
TEMPLATE_SHEETS_CHANGED: Final = "기준 템플릿 시트 순서가 바뀜"
COMPARISONS_REQUIRED: Final = "비교 물품 2개가 필요함"
INVALID_SLOT: Final = "비교 슬롯이 올바르지 않음"


@final
class EstimateExporter:
    """Export one persisted draft through targeted OOXML cell updates."""

    def __init__(
        self,
        database: Path,
        template: Path = DEFAULT_TEMPLATE,
        manifest: Path = DEFAULT_MANIFEST,
    ) -> None:
        """Bind the shared DB and immutable template assets."""
        self.database = database
        self.template = template
        self.manifest = TemplateManifest.model_validate_json(manifest.read_bytes())

    def export(self, estimate_id: str, destination: Path) -> Path:
        """Write one validated draft to a new workbook path."""
        if destination.resolve() == self.template.resolve():
            raise EstimateExportError(TEMPLATE_OVERWRITE)
        self._validate_template()
        draft = EstimateStore(self.database).get_draft(estimate_id)
        comparisons = _read_comparisons(self.database, draft)
        _validate_comparisons(draft, comparisons)
        if draft.template_sha256 != self.manifest.template_sha256:
            raise EstimateExportError(TEMPLATE_HASH_CHANGED)
        with ZipFile(self.template) as archive:
            entries = {name: archive.read(name) for name in archive.namelist()}
        write_draft(entries, self.manifest, draft, comparisons)
        fallback = NO_IMAGE.read_bytes()
        for slot in self.manifest.image_slots:
            if slot.line_index < len(draft.lines):
                entries[slot.media_path] = fallback
        destination.parent.mkdir(parents=True, exist_ok=True)
        with ZipFile(destination, "w", ZIP_DEFLATED, allowZip64=True) as archive:
            for name, data in entries.items():
                archive.writestr(name, data)
        return destination

    def _validate_template(self) -> None:
        actual = hashlib.sha256(self.template.read_bytes()).hexdigest()
        if actual != self.manifest.template_sha256:
            raise EstimateExportError(TEMPLATE_HASH_CHANGED)
        workbook = load_workbook(
            self.template,
            read_only=True,
            data_only=False,
            keep_links=True,
        )
        try:
            if workbook.sheetnames != self.manifest.sheet_names:
                raise EstimateExportError(TEMPLATE_SHEETS_CHANGED)
        finally:
            workbook.close()


def _read_comparisons(
    database: Path,
    draft: EstimateDraft,
) -> dict[str, tuple[ComparisonSnapshot, ...]]:
    result: dict[str, tuple[ComparisonSnapshot, ...]] = {}
    with connect(database) as connection:
        for line in draft.lines:
            rows = query(
                connection,
                """
                SELECT slot, product_id, company_snapshot, spec_snapshot,
                price_won_snapshot FROM estimate_comparisons
                WHERE estimate_line_id = ? ORDER BY slot
                """,
                (line.id,),
            ).fetchall()
            result[line.id] = tuple(
                ComparisonSnapshot(
                    slot=_slot(as_text(row[0])),
                    product_id=as_text(row[1]),
                    company=as_text(row[2]),
                    spec=as_text(row[3]),
                    price_won=as_int(row[4]),
                )
                for row in rows
            )
    return result


def _validate_comparisons(
    draft: EstimateDraft,
    comparisons: dict[str, tuple[ComparisonSnapshot, ...]],
) -> None:
    for line in draft.lines:
        values = comparisons[line.id]
        if tuple(item.slot for item in values) != ("A", "B", "C"):
            raise EstimateExportError(COMPARISONS_REQUIRED)
        selected = values[0]
        if (
            selected.product_id != line.product_id
            or selected.company != line.company_snapshot
            or selected.spec != line.spec_snapshot
            or selected.price_won != line.unit_price_won_snapshot
        ):
            raise EstimateExportError(COMPARISONS_REQUIRED)


def _slot(value: str) -> Literal["A", "B", "C"]:
    match value:
        case "A":
            return "A"
        case "B":
            return "B"
        case "C":
            return "C"
        case _:
            raise EstimateExportError(INVALID_SLOT)
