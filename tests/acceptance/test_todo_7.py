"""Todo 7 delivery-role and curated-relation acceptance contract."""

from __future__ import annotations

import hashlib
import shutil
from contextlib import closing
from dataclasses import replace
from pathlib import Path
from typing import Literal

import pytest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from g2b_compare.importers import workbook_relations
from g2b_compare.importers.workbook_relations import (
    PINNED_FILENAME,
    PINNED_RELATION_MANIFEST,
    RelationImportError,
    RelationImportFailure,
    import_workbook_relations,
)
from g2b_compare.sources.delivery_detail import (
    DeliveryCandidate,
    parse_delivery_candidates,
)

PINNED_WORKBOOK = Path("dataset") / PINNED_FILENAME
OTHER_WORKBOOKS = (
    Path("dataset") / "순천 향교 CCTV 구매 설치 - 내역서(관급)(0706수정).xlsx",
    Path("dataset")
    / "전남 광양시 아트케이션 관광스테이 확충사업 CCTV 설비 - 내역서(관급)(최종).xlsx",
)
type CellDrift = Literal[
    "header-drift",
    "formula-id",
    "merged-id",
    "zero-import",
    "missing-parent",
    "missing-child",
    "self-link",
    "bad-id",
    "duplicate-row",
]
_SIMPLE_MUTATIONS: dict[CellDrift, tuple[str, str | None]] = {
    "header-drift": ("B10", "drift"),
    "formula-id": ("N11", "=24684677"),
    "missing-parent": ("N9", None),
    "missing-child": ("N11", None),
    "self-link": ("N11", "24684676"),
    "bad-id": ("N11", "not-an-id"),
}


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _mutate(scenario: CellDrift, worksheet: Worksheet) -> None:
    mutation = _SIMPLE_MUTATIONS.get(scenario)
    if mutation is not None:
        coordinate, value = mutation
        worksheet[coordinate] = value
        return
    if scenario == "merged-id":
        worksheet.merge_cells("N11:O11")
        return
    if scenario == "duplicate-row":
        worksheet["N12"] = "23842678"
        return
    for row in range(11, 23):
        worksheet[f"N{row}"] = None


def _changed_workbook(tmp_path: Path, scenario: CellDrift) -> Path:
    target = tmp_path / PINNED_FILENAME
    _ = shutil.copyfile(PINNED_WORKBOOK, target)
    with closing(load_workbook(target, read_only=False, data_only=False)) as workbook:
        worksheet = workbook[PINNED_RELATION_MANIFEST.sheet_name]
        assert isinstance(worksheet, Worksheet)
        _mutate(scenario, worksheet)
        workbook.save(target)
    return target


def test_happy_replay_is_immutable_and_delivery_roles_remain_contextual() -> None:
    # Given
    source_sha = _sha256(PINNED_WORKBOOK)
    delivery = (
        DeliveryCandidate(1, "R1", "1", "", "24684676", "대표품목"),
        DeliveryCandidate(2, "R1", "2", "", "24684677", "별도구매선택품목"),
        DeliveryCandidate(3, "R1", "3", "", "24684678", "동시구매품목"),
    )

    # When
    first = import_workbook_relations(PINNED_WORKBOOK)
    replay = import_workbook_relations(PINNED_WORKBOOK)
    events = parse_delivery_candidates(delivery, "2026-07-16T00:00:00Z")

    # Then
    assert first == replay
    assert first.snapshot == replay.snapshot
    assert _sha256(PINNED_WORKBOOK) == source_sha
    assert tuple(event.role_raw for event in events.events) == (
        "대표품목",
        "별도구매선택품목",
        "동시구매품목",
    )
    assert events.relations == ()


def test_failure_missing_delivery_source_key_and_cooccurrence() -> None:
    # Given
    candidates = (
        DeliveryCandidate(1, None, "1", "", "24684676", "대표품목"),
        DeliveryCandidate(2, "R1", "2", "", "24684677", "별도구매선택품목"),
    )

    # When
    result = parse_delivery_candidates(candidates, "2026-07-16T00:00:00Z")

    # Then
    assert tuple(item.source_ordinal for item in result.quarantined) == (1,)
    assert tuple(event.product_id for event in result.events) == ("24684677",)
    assert result.relations == ()


def test_failure_wrong_workbook_and_comparison_columns_are_not_relations() -> None:
    results = tuple(import_workbook_relations(path) for path in OTHER_WORKBOOKS)

    assert all(result.relations == () for result in results)
    assert all(result.snapshot is None for result in results)


@pytest.mark.parametrize(
    ("scenario", "expected"),
    [
        ("header-drift", RelationImportFailure.HEADER_DRIFT),
        ("formula-id", RelationImportFailure.FORMULA_ID),
        ("merged-id", RelationImportFailure.MERGED_ID),
        ("zero-import", RelationImportFailure.ZERO_IMPORT),
        ("missing-parent", RelationImportFailure.MISSING_PARENT),
        ("missing-child", RelationImportFailure.MISSING_CHILD),
        ("self-link", RelationImportFailure.SELF_LINK),
        ("bad-id", RelationImportFailure.BAD_ID),
        ("duplicate-row", RelationImportFailure.DUPLICATE_ROW),
    ],
    ids=(
        "header-drift",
        "formula-id",
        "merged-id",
        "zero-import",
        "missing-parent",
        "missing-child",
        "self-link",
        "bad-id",
        "duplicate-row",
    ),
)
def test_failure_target_cell_drift_fails_closed(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    scenario: CellDrift,
    expected: RelationImportFailure,
) -> None:
    # Given
    changed = _changed_workbook(tmp_path, scenario)
    manifest = replace(PINNED_RELATION_MANIFEST, source_sha256=_sha256(changed))
    monkeypatch.setattr(workbook_relations, "PINNED_RELATION_MANIFEST", manifest)

    # When / Then
    with pytest.raises(RelationImportError) as captured:
        _ = import_workbook_relations(changed, manifest)
    assert captured.value.failure is expected


def test_failure_source_hash_mismatch_is_not_a_zero_import(tmp_path: Path) -> None:
    # Given
    changed = _changed_workbook(tmp_path, "header-drift")

    # When / Then
    with pytest.raises(RelationImportError) as captured:
        _ = import_workbook_relations(changed)
    assert captured.value.failure is RelationImportFailure.SOURCE_HASH_MISMATCH


def test_failure_unexpected_relation_count_fails_closed() -> None:
    # Given
    manifest = replace(PINNED_RELATION_MANIFEST, child_coordinates=("N11",))

    # When / Then
    with pytest.raises(RelationImportError) as captured:
        _ = import_workbook_relations(PINNED_WORKBOOK, manifest)
    assert captured.value.failure is RelationImportFailure.MANIFEST_DRIFT


def test_failure_manifest_coordinate_order_substitution_fails_closed() -> None:
    # Given
    manifest = replace(
        PINNED_RELATION_MANIFEST,
        child_coordinates=tuple(reversed(PINNED_RELATION_MANIFEST.child_coordinates)),
    )

    # When / Then
    with pytest.raises(RelationImportError) as captured:
        _ = import_workbook_relations(PINNED_WORKBOOK, manifest)
    assert captured.value.failure is RelationImportFailure.MANIFEST_DRIFT


def test_failure_unbound_options_remain_quarantined() -> None:
    result = import_workbook_relations(PINNED_WORKBOOK)

    relation_ids = {relation.child_id for relation in result.relations}
    quarantine_ids = {item.product_id for item in result.quarantined}
    assert len(quarantine_ids) == 3
    assert relation_ids.isdisjoint(quarantine_ids)
