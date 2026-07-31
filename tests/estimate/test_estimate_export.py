from __future__ import annotations

import hashlib
import json
import sqlite3
from decimal import Decimal
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from g2b_compare import services

ASSETS = Path(services.__file__).resolve().parents[1] / "assets"
TEMPLATE = ASSETS / "estimate-template-v1.xlsx"
MANIFEST = ASSETS / "estimate-template-v1.json"


def _line(index: int) -> services.EstimateLineInput:
    return services.EstimateLineInput(
        line_kind="main",
        product_id=f"25{index:06d}",
        parent_product_id=None,
        relation_id=None,
        offer_operation="getMASCntrctPrdctInfoList",
        offer_key=f"offer-{index}",
        item_name_snapshot=f"영상감시장치 {index}",
        spec_snapshot=f"800만화소 {index}",
        company_snapshot=f"A 공급사 {index}",
        unit_snapshot="조",
        unit_price_won_snapshot=1_000_000 + index,
        quantity=Decimal(index),
    )


def _draft_with_lines(database: Path, count: int) -> services.EstimateDraft:
    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
    store = services.EstimateStore(database)
    draft = store.create_draft(
        "순천 향교 CCTV 구매 설치",
        str(manifest["template_sha256"]),
    )
    lines = tuple(
        store.add_line(draft.id, _line(index)) for index in range(1, count + 1)
    )
    with sqlite3.connect(database) as connection:
        for line in lines:
            connection.executemany(
                "INSERT INTO estimate_comparisons VALUES (?, ?, ?, ?, ?, ?, ?)",
                (
                    (
                        line.id,
                        slot,
                        (
                            line.product_id
                            if slot == "A"
                            else f"26{line.line_no:05d}{slot_index}"
                        ),
                        None,
                        f"{slot} 공급사",
                        f"비교 규격 {slot}",
                        price,
                    )
                    for slot_index, (slot, price) in enumerate(
                        (
                            ("A", line.unit_price_won_snapshot),
                            ("B", 900_000),
                            ("C", 1_100_000),
                        ),
                        start=1,
                    )
                ),
            )
    return store.get_draft(draft.id)


@pytest.mark.parametrize("line_count", [1, 9])
def test_export_preserves_template_and_maps_one_or_nine_lines(
    tmp_path: Path,
    line_count: int,
) -> None:
    database = tmp_path / "g2b.sqlite3"
    draft = _draft_with_lines(database, line_count)
    before_sha = hashlib.sha256(TEMPLATE.read_bytes()).hexdigest()
    destination = tmp_path / f"estimate-{line_count}.xlsx"

    exported = services.EstimateExporter(database).export(draft.id, destination)

    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
    workbook = load_workbook(exported, data_only=False, keep_links=True)
    quantity_sheet = workbook[str(manifest["quantity_rows"]["sheet"])]
    price_sheet = workbook[str(manifest["price_rows"]["sheet"])]
    title_sheet = workbook[str(manifest["title_cell"]["sheet"])]
    procurement_sheet = workbook[manifest["sheet_names"][11]]

    assert workbook.sheetnames == manifest["sheet_names"]
    assert title_sheet["A5"].value == "순천 향교 CCTV 구매 설치"
    assert quantity_sheet["A8"].value == "1-1"
    assert quantity_sheet["B8"].value == "영상감시장치 1"
    assert quantity_sheet["F8"].value == 1
    assert price_sheet["F5"].value == "A 공급사"
    assert price_sheet["I5"].value == 1_000_001
    assert price_sheet["E5"].value == "=MIN(I5,M5,Q5)"
    assert procurement_sheet["L19"].value == "=SUM(L5:L17)"
    assert procurement_sheet["L20"].value == "=L19*0.0054"
    assert procurement_sheet["L21"].value == "=ROUNDUP(SUM(L19:L20),-3)"
    assert quantity_sheet[f"B{7 + line_count}"].value == f"영상감시장치 {line_count}"
    if line_count == 1:
        assert quantity_sheet["B9"].value is None
        assert price_sheet["F6"].value is None
    else:
        assert quantity_sheet["B16"].value == "영상감시장치 9"
        assert price_sheet["F13"].value == "A 공급사"

    with ZipFile(exported) as archive:
        drawing = archive.read("xl/drawings/drawing5.xml")
        assert drawing.count(b"rIdEstimate") == 23
        for slot in manifest["image_slots"]:
            assert archive.read(slot["media_path"]).startswith(b"\x89PNG")

    assert hashlib.sha256(TEMPLATE.read_bytes()).hexdigest() == before_sha


def test_export_rejects_line_without_two_comparison_products(tmp_path: Path) -> None:
    database = tmp_path / "g2b.sqlite3"
    store = services.EstimateStore(database)
    draft = store.create_draft("비교 부족", "a" * 64)
    store.add_line(draft.id, _line(1))

    with pytest.raises(services.EstimateExportError, match="비교 물품 2개가 필요함"):
        services.EstimateExporter(database).export(
            draft.id,
            tmp_path / "blocked.xlsx",
        )


def test_export_rejects_draft_pinned_to_a_different_template(tmp_path: Path) -> None:
    database = tmp_path / "g2b.sqlite3"
    draft = _draft_with_lines(database, 1)
    with sqlite3.connect(database) as connection:
        connection.execute(
            "UPDATE estimate_drafts SET template_sha256 = ? WHERE id = ?",
            ("a" * 64, draft.id),
        )

    with pytest.raises(
        services.EstimateExportError,
        match="기준 템플릿 해시가 일치하지 않음",
    ):
        services.EstimateExporter(database).export(
            draft.id,
            tmp_path / "blocked.xlsx",
        )


def test_export_accepts_koreanet_baseline_for_other_selected_product(
    tmp_path: Path,
) -> None:
    database = tmp_path / "g2b.sqlite3"
    draft = _draft_with_lines(database, 1)
    with sqlite3.connect(database) as connection:
        connection.execute(
            """
            UPDATE estimate_comparisons
            SET product_id = ?, company_snapshot = ?
            WHERE estimate_line_id = ? AND slot = 'A'
            """,
            ("25454886", "주식회사 코리아넷", draft.lines[0].id),
        )

    exported = services.EstimateExporter(database).export(
        draft.id,
        tmp_path / "koreanet-baseline.xlsx",
    )

    assert exported.is_file()
