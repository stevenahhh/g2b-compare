"""Task 13 production-bundle browser journeys."""

from __future__ import annotations

import asyncio
import io
import json
import re
import threading
import time
from collections import Counter, defaultdict
from collections.abc import Iterator
from pathlib import Path
from tempfile import TemporaryDirectory
from urllib.parse import parse_qs, urlparse

import pytest
import uvicorn
from anyio import Path as AsyncPath
from openpyxl import Workbook, load_workbook
from playwright.async_api import BrowserContext, Page, Route, async_playwright, expect

from g2b_compare.web import app as app_module

DIST = Path("src/g2b_compare/web/frontend_dist")
EVIDENCE = Path(".omo/evidence/svelte-spa-single-user")
REFRESH_EVIDENCE = Path(".omo/evidence/document-ui-refresh")


@pytest.fixture(scope="module")
def spa_url() -> Iterator[str]:
    assert (DIST / "index.html").is_file(), "production SPA bundle is required"
    with TemporaryDirectory() as temporary:
        home = Path(temporary)
        with pytest.MonkeyPatch.context() as monkeypatch:
            monkeypatch.setattr(app_module, "FRONTEND_DIST", DIST)
            monkeypatch.setenv("G2B_SERVE_SPA", "1")
            app = app_module.create_app(database=home / "g2b.sqlite3", home=home)
            server = uvicorn.Server(
                uvicorn.Config(app, host="127.0.0.1", port=0, log_level="warning", access_log=False)
            )
            thread = threading.Thread(target=server.run, daemon=True)
            thread.start()
            for _ in range(100):
                if server.started:
                    break
                time.sleep(0.02)
            else:
                raise AssertionError("FastAPI SPA server did not start")
            socket = next(iter(server.servers)).sockets[0]
            try:
                yield f"http://127.0.0.1:{socket.getsockname()[1]}"
            finally:
                server.should_exit = True
                thread.join(timeout=5)
                assert not thread.is_alive()


def _product(number: int = 1) -> dict[str, object]:
    return {
        "product_id": f"{number:08d}",
        "name": "실외형 네트워크 카메라" if number == 1 else f"보조 카메라 {number}",
        "spec": "4K · IP66",
        "unit": "대",
        "price_won": 1_250_000,
        "company_name": "주식회사 화면",
        "contract_method": "다수공급자계약",
        "delivery_condition": "현장설치도",
        "delivery_days": "14",
        "contract_end_date": "2027-12-31",
        "detail_url": "https://example.test/detail/main",
        "g2b_url": "https://example.test/g2b/main",
        "image_url": "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='92' height='92'%3E%3Crect width='92' height='92' fill='%230369a1'/%3E%3C/svg%3E",
        "attributes": [{"name": "해상도", "value": "3840×2160", "unit": "px"}, {"name": "방수", "value": "IP66", "unit": ""}],
    }


def _option() -> dict[str, object]:
    return {
        "parent_product_id": "00000001",
        "parent_name": "실외형 네트워크 카메라",
        "relation_id": "REL-PTZ-01",
        "relation_kind": "component",
        "product_id": "00000002",
        "name": "PTZ 추가 선택품목",
        "spec": "30배 광학 줌",
        "unit": "식",
        "price_won": 350_000,
        "company_name": "주식회사 옵션",
        "detail_url": "https://example.test/detail/option",
        "g2b_url": "https://example.test/g2b/option",
        "image_url": "",
        "attributes": [{"name": "광학줌", "value": "30", "unit": "배"}],
    }


def _comparison(line: dict[str, object], slot: str) -> dict[str, object]:
    suffix = {"A": "알파", "B": "베타", "C": "감마"}[slot]
    return {
        "slot": slot,
        "product_id": line["product_id"],
        "relation_id": line.get("relation_id"),
        "company_snapshot": f"비교{suffix}",
        "spec_snapshot": f"{line['spec_snapshot']} {suffix}",
        "price_won_snapshot": int(line["unit_price_won_snapshot"]) + {"A": 0, "B": 1000, "C": 2000}[slot],
        "attributes": [{"name": "비교속성", "value": suffix, "unit": ""}],
    }


def _remote_estimate(estimate_id: str, payload: dict[str, object]) -> dict[str, object]:
    lines = []
    for position, source in enumerate(payload["lines"], start=1):
        line = dict(source)
        line.update(
            {
                "line_no": position,
                "quantity": str(line["quantity"]),
                "attributes": _product()["attributes"] if line["line_kind"] == "main" else _option()["attributes"],
                "comparisons": [_comparison(line, slot) for slot in ("A", "B", "C")],
            }
        )
        lines.append(line)
    return {
        "id": estimate_id,
        "title": payload["title"],
        "created_at": "2026-07-22T10:00:00+00:00",
        "updated_at": "2026-07-22T10:00:01+00:00",
        "lines": lines,
        "export_ready": True,
    }


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "관급내역"
    sheet.append(["품명", "수량", "금액"])
    sheet.append(["실외형 네트워크 카메라", 1, 1_250_000])
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


async def _activate_service_worker(page: Page) -> None:
    await page.evaluate("() => navigator.serviceWorker.ready")
    await page.reload(wait_until="networkidle")
    assert await page.evaluate("() => Boolean(navigator.serviceWorker.controller)")




async def _install_contract_routes(
    page: Page,
    estimates: dict[str, dict[str, object]],
    counters: dict[str, Counter[str]],
    offline: dict[str, bool],
    *,
    total_count: int = 90,
    refresh_gate: tuple[asyncio.Event, asyncio.Event] | None = None,
) -> asyncio.Queue[int]:
    sync_events: asyncio.Queue[int] = asyncio.Queue()
    async def api(route: Route) -> None:
        request = route.request
        path = urlparse(request.url).path
        if offline["enabled"]:
            await route.abort("failed")
            return
        counters[request.method][path] += 1
        if path == "/api/catalog/products" and request.method == "GET":
            query = parse_qs(urlparse(request.url).query)
            page_number = int(query.get("page", ["1"])[0])
            page_size = int(query.get("page_size", ["30"])[0])
            page_count = max(1, -(-total_count // page_size))
            first = (page_number - 1) * page_size + 1
            items = [_product(number) for number in range(first, min(first + page_size, total_count + 1))]
            await route.fulfill(json={"items": items, "page": page_number, "page_count": page_count, "total_count": total_count})
            return
        if path == "/api/catalog/relations" and request.method == "GET":
            query = parse_qs(urlparse(request.url).query)
            items = [_option()] if query.get("category") == ["selection"] else []
            await route.fulfill(
                json={
                    "items": items,
                    "page": 1,
                    "page_count": 1,
                    "total_count": len(items),
                }
            )
            return
        if (
            path.startswith("/api/catalog/products/")
            and path.endswith("/options")
            and request.method == "GET"
        ):
            await route.fulfill(
                json={
                    "items": [_option()],
                    "page": 1,
                    "page_count": 1,
                    "total_count": 1,
                }
            )
            return
        if path == "/api/estimates" and request.method == "GET":
            await route.fulfill(json=[{"id": estimate_id, "title": document["title"], "updated_at": "2026-07-22T10:00:01+00:00", "line_count": len(document["lines"])} for estimate_id, document in estimates.items()])
            return
        if path.endswith("/refresh-comparisons") and request.method == "POST":
            estimate_id = path.split("/")[-2]
            if refresh_gate is not None:
                started, release = refresh_gate
                started.set()
                await release.wait()
            refreshed = _remote_estimate(estimate_id, estimates[estimate_id])
            for line in refreshed["lines"]:
                line["comparisons"][0]["company_snapshot"] = "새 비교군"
            await route.fulfill(status=200, json=refreshed)
            return
        if path.startswith("/api/estimates/"):
            estimate_id = path.rsplit("/", 1)[-1]
            if request.method == "PUT":
                estimates[estimate_id] = json.loads(request.post_data or "{}")
                await route.fulfill(
                    status=200,
                    json=_remote_estimate(estimate_id, estimates[estimate_id]),
                )
                sync_events.put_nowait(len(estimates[estimate_id]["lines"]))
                return
            if request.method == "DELETE":
                estimates.pop(estimate_id, None)
                await route.fulfill(status=204, body="")
                return
            if request.method == "GET" and estimate_id in estimates:
                await route.fulfill(json=_remote_estimate(estimate_id, estimates[estimate_id]))
                return
        await route.fulfill(status=404, json={"detail": "unhandled test contract"})

    await page.route("**/api/**", api)
    return sync_events


async def _read_estimates(page: Page) -> list[dict[str, object]]:
    return await page.evaluate(
        """() => new Promise((resolve, reject) => {
          const open = indexedDB.open('g2b-spa', 1);
          open.onerror = () => reject(open.error);
          open.onsuccess = () => {
            const tx = open.result.transaction('estimates', 'readonly');
            const get = tx.objectStore('estimates').getAll();
            get.onerror = () => reject(get.error);
            get.onsuccess = () => resolve(get.result);
          };
        })"""
    )


async def _wait_for_server_lines(
    sync_events: asyncio.Queue[int],
    line_count: int,
) -> None:
    while True:
        synced_lines = await asyncio.wait_for(sync_events.get(), timeout=5)
        if synced_lines >= line_count:
            return


async def _capture_responsive(
    page: Page,
    stem: str,
    evidence: Path = EVIDENCE,
) -> None:
    for width, height in ((375, 812), (768, 900), (1280, 900)):
        await page.set_viewport_size({"width": width, "height": height})
        await page.mouse.move(1, 1)
        await page.evaluate(
            """() => Promise.all(
              document.getAnimations()
                .filter((animation) =>
                  Number.isFinite(animation.effect?.getTiming().iterations ?? 1)
                )
                .map((animation) => animation.finished.catch(() => undefined))
            )"""
        )
        await page.screenshot(
            path=evidence / f"{stem}-{width}x{height}.png",
            full_page=False,
        )



@pytest.mark.asyncio
async def test_svelte_spa_catalog_to_comparison_journey(spa_url: str) -> None:
    EVIDENCE.mkdir(parents=True, exist_ok=True)
    estimates: dict[str, dict[str, object]] = {}
    counters: dict[str, Counter[str]] = defaultdict(Counter)
    offline = {"enabled": False}
    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch()
        context = await browser.new_context(viewport={"width": 1440, "height": 900}, accept_downloads=True)
        page = await context.new_page()
        sync_events = await _install_contract_routes(
            page,
            estimates,
            counters,
            offline,
        )
        _ = await page.goto(spa_url, wait_until="networkidle")
        await expect(page.get_by_text("실외형 네트워크 카메라").first).to_be_visible()
        assert await page.locator(".catalog-card").count() > 0
        assert await page.locator(".catalog-card").count() <= 60
        assert await page.locator(".g2b-link").first.evaluate("node => getComputedStyle(node).width") == "150px"
        catalog = page.locator(".catalog-scroll")
        for page_number in (2, 3):
            async with page.expect_response(
                lambda response, expected=page_number: (
                    urlparse(response.url).path == "/api/catalog/products"
                    and parse_qs(urlparse(response.url).query).get("page")
                    == [str(expected)]
                )
            ):
                await catalog.evaluate(
                    "element => { element.scrollTop = element.scrollHeight; "
                    "element.dispatchEvent(new Event('scroll')); }"
                )
        assert counters["GET"]["/api/catalog/products"] == 3
        assert await page.locator(".catalog-card").count() <= 60
        logical_rows = await page.evaluate(
            """() => {
              const grid = document.querySelector(".catalog-grid");
              const rowHeight = parseFloat(
                grid.style.getPropertyValue("--catalog-row")
              );
              const spacers = [...grid.querySelectorAll(".virtual-spacer")];
              const hidden = spacers.reduce(
                (total, node) => total + parseFloat(node.style.height || "0") / rowHeight,
                0
              );
              return hidden + grid.querySelectorAll(".catalog-card").length;
            }"""
        )
        assert logical_rows >= 90
        await catalog.evaluate(
            "element => { element.scrollTop = 0; "
            "element.dispatchEvent(new Event('scroll')); }"
        )
        await expect(page.get_by_text("실외형 네트워크 카메라").first).to_be_visible()
        await page.get_by_role("button", name="실외형 네트워크 카메라").click()
        await expect(page.get_by_text("PTZ 추가 선택품목")).to_be_visible()
        await expect(page.get_by_text("해상도").first).to_be_visible()
        await expect(page.get_by_text("30배 광학 줌")).to_be_visible()
        await page.locator(".catalog-card").first.get_by_role(
            "button", name="리스트에 추가"
        ).click()
        await _wait_for_server_lines(sync_events, 1)
        await page.get_by_role("button", name="문서 작성").click()
        await expect(page.locator(".estimate-summary")).to_have_count(1)
        await page.locator(".estimate-summary__open").click()
        title_button = page.locator(".page-title-edit")
        original_title = await title_button.text_content()
        await title_button.click()
        title_input = page.get_by_role("textbox", name="문서 제목")
        await expect(title_input).to_be_focused()
        await title_input.fill("취소할 제목")
        await title_input.press("Escape")
        await expect(title_button).to_have_text(original_title or "")

        await page.get_by_role("button", name="내역 추가").click()
        product_search = page.locator("#document-product-search")
        await expect(product_search).to_be_focused()
        await page.locator("#document-product-sort").focus()
        await page.locator("#document-product-sort").press("Escape")
        await expect(page.locator(".document-search-overlay")).not_to_have_class(
            re.compile(r"\bis-open\b")
        )
        await expect(product_search).to_be_focused()
        await page.get_by_role("button", name="내역 추가").click()
        await expect(
            page.locator(".document-search-overlay .catalog-card").first
        ).to_be_visible()
        await page.locator(
            ".document-search-overlay .catalog-card__select"
        ).first.click()
        await expect(page.locator(".document-option-panel")).to_be_visible()
        await page.locator(".document-option-row").first.get_by_role(
            "button", name="리스트에 추가"
        ).click()
        await _wait_for_server_lines(sync_events, 2)
        await expect(page.locator(".document-table tbody tr")).to_have_count(2)
        for heading in ("적용회사(A사)", "B사", "C사"):
            await expect(
                page.get_by_role("columnheader", name=heading, exact=True)
            ).to_be_visible()
        document_table = page.locator(".document-table")
        await expect(document_table).to_contain_text("3840×2160, IP66")
        await expect(document_table).to_contain_text("30배 광학 줌")
        tooltip_trigger = page.locator(".spec-tooltip-trigger").first
        await tooltip_trigger.focus()
        await expect(page.get_by_role("tooltip")).to_be_visible()
        await tooltip_trigger.press("Escape")
        await expect(page.get_by_role("tooltip")).to_have_count(0)
        await page.route(
            "**/estimates/*/export.xlsx",
            lambda route: route.fulfill(
                status=200,
                body=_workbook_bytes(),
                headers={
                    "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "content-disposition": 'attachment; filename="estimate.xlsx"',
                },
            ),
        )
        async with page.expect_download() as xlsx_download:
            await page.get_by_role("link", name="XLSX 내려받기").click()
        xlsx = await xlsx_download.value
        assert xlsx.suggested_filename.endswith(".xlsx")
        workbook = load_workbook(io.BytesIO(Path(await xlsx.path()).read_bytes()), data_only=True)
        assert workbook["관급내역"]["A2"].value == "실외형 네트워크 카메라"
        assert workbook["관급내역"]["B2"].value == 1
        await _capture_responsive(page, "catalog-comparison")
        (EVIDENCE / "catalog-comparison-counters.json").write_text(json.dumps({method: dict(count) for method, count in counters.items()}, ensure_ascii=False, indent=2), encoding="utf-8")
        await context.close()
        await browser.close()


async def _read_document_ui_contract(page: Page) -> dict[str, int]:
    return {
        "tsv": await page.get_by_role(
            "button", name="TSV 내려받기", exact=True
        ).count(),
        "refresh": await page.get_by_role(
            "button", name="비교군 새로고침", exact=True
        ).count(),
        "comparison_region": await page.get_by_role(
            "region", name="단가 비교표", exact=True
        ).count(),
        "comparison_heading": await page.get_by_role(
            "heading", name="비교군 목록", exact=True
        ).count(),
        "comparison_copy": await page.get_by_text(
            "현재 단가를 기준으로 A·B·C 후보를 문서에 고정함",
            exact=True,
        ).count(),
        "scroll_copy": await page.get_by_text(
            "좌우로 스크롤해 B·C 후보까지 확인",
            exact=True,
        ).count(),
        "refresh_in_actions": await page.locator(
            ".page-actions .comparison-refresh"
        ).count(),
        "refresh_in_table": await page.locator(
            ".document-sheet .comparison-refresh"
        ).count(),
    }


async def _read_document_list_contract(page: Page) -> dict[str, int]:
    return {
        "home": await page.get_by_role("button", name="홈", exact=True).count(),
        "new_document": await page.get_by_role(
            "button", name="새 문서", exact=True
        ).count(),
        "legacy_start": await page.get_by_role(
            "button", name="새 내역 시작", exact=True
        ).count(),
    }


async def _capture_mobile_refreshed_candidate(page: Page) -> None:
    await page.set_viewport_size({"width": 375, "height": 812})
    await page.locator(".document-table-wrap").evaluate(
        "element => { element.scrollLeft = 500; }"
    )
    refreshed_company = page.get_by_role(
        "cell",
        name="새 비교군",
        exact=True,
    )
    assert await refreshed_company.evaluate(
        """element => {
          const bounds = element.getBoundingClientRect();
          return bounds.left >= 0 && bounds.right <= innerWidth;
        }"""
    )
    _ = await page.screenshot(
        path=REFRESH_EVIDENCE / "editor-refreshed-candidate-375x812.png",
        full_page=False,
    )


@pytest.mark.asyncio
async def test_svelte_spa_document_ui_refresh_contract(spa_url: str) -> None:
    # Given: the live SPA uses a deterministic estimate API contract.
    await AsyncPath(REFRESH_EVIDENCE).mkdir(parents=True, exist_ok=True)
    estimates: dict[str, dict[str, object]] = {}
    counters: dict[str, Counter[str]] = defaultdict(Counter)
    offline = {"enabled": False}
    refresh_started = asyncio.Event()
    refresh_release = asyncio.Event()
    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch()
        context = await browser.new_context(viewport={"width": 1280, "height": 900})
        page = await context.new_page()
        await page.add_init_script(
            """class ContractEventSource {
              static current;
              constructor() {
                this.listeners = new Map();
                ContractEventSource.current = this;
              }
              addEventListener(name, listener) {
                const listeners = this.listeners.get(name) ?? [];
                listeners.push(listener);
                this.listeners.set(name, listeners);
              }
              close() {}
              emit(name, data) {
                for (const listener of this.listeners.get(name) ?? []) {
                  listener(new MessageEvent(name, {
                    data: JSON.stringify(data)
                  }));
                }
              }
            }
            window.EventSource = ContractEventSource;
            window.__emitEstimateEvent = (name, id) =>
              ContractEventSource.current?.emit(name, { id });"""
        )
        sync_events = await _install_contract_routes(
            page,
            estimates,
            counters,
            offline,
            refresh_gate=(refresh_started, refresh_release),
        )
        _ = await page.goto(spa_url, wait_until="networkidle")
        await page.get_by_role("button", name="문서 작성").click()
        await page.get_by_role("button", name="물품 검색", exact=True).click()
        await expect(page.locator(".catalog-card").first).to_be_visible()
        await page.locator(".catalog-card").first.get_by_role(
            "button", name="리스트에 추가"
        ).click()
        await _wait_for_server_lines(sync_events, 1)
        await page.get_by_role("button", name="문서 작성").click()
        await expect(page.locator(".estimate-summary")).to_have_count(1)
        list_contract = await _read_document_list_contract(page)
        await _capture_responsive(page, "list", REFRESH_EVIDENCE)
        await page.locator(".estimate-summary__open").click()
        refresh_button = page.locator(".comparison-refresh")
        await expect(refresh_button).to_be_visible()
        await expect(refresh_button).to_be_enabled()
        await expect(refresh_button).to_have_accessible_name("비교군 새로고침")
        editor_contract = await _read_document_ui_contract(page)
        await _capture_responsive(page, "editor", REFRESH_EVIDENCE)

        # When: the requested document-list and editor surface is inspected.
        observed = {**list_contract, **editor_contract}

        # Then: legacy/TSV controls are absent and refresh owns a named region.
        assert observed == {
            "home": 0,
            "new_document": 1,
            "legacy_start": 0,
            "tsv": 0,
            "refresh": 1,
            "comparison_region": 1,
            "comparison_heading": 0,
            "comparison_copy": 0,
            "scroll_copy": 0,
            "refresh_in_actions": 1,
            "refresh_in_table": 0,
        }
        refresh_path = next(iter(estimates))
        async with page.expect_response(
            lambda response: (
                urlparse(response.url).path
                == f"/api/estimates/{refresh_path}/refresh-comparisons"
                and response.request.method == "POST"
            )
        ):
            _ = await page.evaluate(
                "() => document.querySelector('.comparison-refresh').click()"
            )
            try:
                await asyncio.wait_for(refresh_started.wait(), timeout=5)
                await expect(refresh_button).to_be_disabled()
                await expect(refresh_button).to_have_text("새로고침 중")
                await expect(page.locator(".document-table-wrap")).to_have_attribute(
                    "aria-busy",
                    "true",
                )
                await expect(
                    page.get_by_role("button", name="내역 추가", exact=True)
                ).to_be_disabled()
                await expect(
                    page.locator(".document-table .document-sequence button").first
                ).to_be_disabled()
                await expect(
                    page.locator(
                        ".document-search-overlay .catalog-card__actions button"
                    ).first
                ).to_be_disabled()
                await _capture_responsive(
                    page,
                    "editor-refreshing",
                    REFRESH_EVIDENCE,
                )
            finally:
                _ = refresh_release.set()
        await expect(page.locator(".document-table")).to_contain_text("새 비교군")
        await expect(refresh_button).to_have_text("새로고침 완료")
        assert (
            counters["POST"][f"/api/estimates/{refresh_path}/refresh-comparisons"]
            == 1
        )
        await _capture_responsive(
            page,
            "editor-refreshed",
            REFRESH_EVIDENCE,
        )
        await _capture_mobile_refreshed_candidate(page)
        remote_gets_before_event = counters["GET"][
            f"/api/estimates/{refresh_path}"
        ]
        external_title = "외부에서 갱신된 문서"
        estimates[refresh_path] = {
            **estimates[refresh_path],
            "title": external_title,
        }
        _ = await page.evaluate(
            "id => window.__emitEstimateEvent('estimate-saved', id)",
            refresh_path,
        )
        await expect(page.locator(".document-table")).to_contain_text("비교알파")
        assert (
            counters["GET"][f"/api/estimates/{refresh_path}"]
            == remote_gets_before_event + 1
        )
        await expect(page.get_by_role("heading", level=1)).to_have_text(
            external_title
        )
        estimates.pop(refresh_path)
        _ = await page.evaluate(
            "id => window.__emitEstimateEvent('estimate-deleted', id)",
            refresh_path,
        )
        await expect(page.get_by_role("heading", level=2)).to_have_text(
            "문서를 찾을 수 없음"
        )
        await context.close()
        await browser.close()


@pytest.mark.asyncio
async def test_svelte_spa_warm_cache_offline_latest_state_replay(spa_url: str) -> None:
    EVIDENCE.mkdir(parents=True, exist_ok=True)
    estimates: dict[str, dict[str, object]] = {}
    counters: dict[str, Counter[str]] = defaultdict(Counter)
    offline = {"enabled": False}
    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch()
        context: BrowserContext = await browser.new_context(viewport={"width": 1440, "height": 900})
        page = await context.new_page()
        sync_events = await _install_contract_routes(
            page,
            estimates,
            counters,
            offline,
            total_count=10_000,
        )
        await page.goto(spa_url, wait_until="networkidle")
        await page.get_by_role("button", name="실외형 네트워크 카메라").click()
        await expect(page.get_by_text("PTZ 추가 선택품목")).to_be_visible()
        timing = await page.evaluate(
            """() => new Promise((resolve, reject) => {
              const start = performance.now(); const open = indexedDB.open('g2b-spa', 1);
              open.onerror = () => reject(open.error);
              open.onsuccess = () => {
                const tx = open.result.transaction('catalog_cache', 'readonly');
                const get = tx.objectStore('catalog_cache').get('catalog:v3:주식회사 코리아넷::price_asc:1');
                get.onerror = () => reject(get.error);
                get.onsuccess = () => resolve({milliseconds: performance.now() - start, found: Boolean(get.result?.value)});
              };
            })"""
        )
        assert timing["found"] is True
        assert timing["milliseconds"] < 100
        assert await page.locator(".catalog-card").count() <= 60
        assert await page.locator(".catalog-card").count() < 10_000
        await page.locator(".catalog-card").first.get_by_role(
            "button", name="리스트에 추가"
        ).click()
        await _wait_for_server_lines(sync_events, 1)
        await page.get_by_role("button", name="문서 작성").click()
        await expect(page.locator(".estimate-summary")).to_have_count(1)
        await page.locator(".estimate-summary__open").click()
        estimate_id = (await page.locator(".route-id").text_content()).split(": ", 1)[1]
        for _ in range(100):
            if all(not record["pendingSync"] for record in await _read_estimates(page)):
                break
            await asyncio.sleep(0.02)
        else:
            raise AssertionError("initial estimate sync did not settle")
        await _activate_service_worker(page)
        offline["enabled"] = True
        await context.set_offline(True)
        offline_title = "오프라인 문서"
        await page.locator(".page-title-edit").click()
        title_input = page.get_by_role("textbox", name="문서 제목")
        await title_input.fill(offline_title)
        await title_input.press("Enter")
        for _ in range(100):
            pending = await _read_estimates(page)
            if pending and pending[0]["pendingSync"] and pending[0]["document"]["title"] == offline_title:
                break
            await asyncio.sleep(0.02)
        else:
            raise AssertionError("offline title edit was not persisted")
        await page.get_by_role("button", name="물품 검색", exact=True).click()
        await expect(page.locator(".catalog-card").first).to_be_visible()
        await expect(page.locator(".offline-banner")).to_be_visible()
        await page.get_by_role("button", name="문서 작성").click()
        await page.locator(".estimate-summary__open").click()
        records = await _read_estimates(page)
        assert records[0]["pendingSync"] is True
        assert records[0]["document"]["title"] == offline_title
        await page.get_by_role("button", name="삭제").first.click()
        for _ in range(100):
            tombstone = await _read_estimates(page)
            if tombstone and tombstone[0]["deleted"] and tombstone[0]["pendingSync"]:
                break
            await asyncio.sleep(0.02)
        else:
            raise AssertionError("offline deletion tombstone was not persisted")
        await page.reload(wait_until="domcontentloaded")
        await expect(page.locator(".offline-banner")).to_be_visible()
        records = await _read_estimates(page)
        assert len(records) == 1 and records[0]["deleted"] is True and records[0]["pendingSync"] is True
        await _capture_responsive(page, "warm-cache-offline-pending")
        for _ in range(100):
            failed = await _read_estimates(page)
            if failed and failed[0]["error"]:
                break
            await asyncio.sleep(0.02)
        else:
            raise AssertionError("failed replay did not retain an actionable error")
        counters["PUT"].clear()
        counters["DELETE"].clear()
        offline["enabled"] = False
        await context.set_offline(False)
        await page.get_by_role("button", name="문서 작성", exact=True).click()
        await page.get_by_role("button", name="재시도").click()
        await expect(page.locator(".estimate-summary")).to_have_count(0)
        for _ in range(100):
            if counters["DELETE"][f"/api/estimates/{estimate_id}"] == 1:
                break
            await asyncio.sleep(0.02)
        else:
            raise AssertionError("latest deletion was not replayed")
        assert counters["PUT"][f"/api/estimates/{estimate_id}"] <= 1
        assert counters["DELETE"][f"/api/estimates/{estimate_id}"] <= 1
        assert counters["DELETE"][f"/api/estimates/{estimate_id}"] == 1
        (EVIDENCE / "warm-cache-offline-timing-counters.json").write_text(json.dumps({"warmIndexedDbGetMilliseconds": timing["milliseconds"], "counters": {method: dict(count) for method, count in counters.items()}}, ensure_ascii=False, indent=2), encoding="utf-8")
        await context.close()
        await browser.close()
