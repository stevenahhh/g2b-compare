"""Deterministic facts and smoke cases from the three pinned workbooks."""

from __future__ import annotations

import hashlib
import re
import zipfile
from contextlib import closing
from typing import TYPE_CHECKING, Final

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple

from .models import (
    RELATION_GRAMMAR,
    FixtureBundle,
    FixtureError,
    Manifest,
    SheetFacts,
    SmokeFixture,
    WorkbookFacts,
    WorkbookTotals,
)

if TYPE_CHECKING:
    from pathlib import Path

    from openpyxl.cell.cell import Cell, MergedCell
    from openpyxl.worksheet.worksheet import Worksheet

EXPECTED_SHA256: Final = frozenset(
    {
        "445012e259ab5318a1d52468cce93ee28a55a8bcb467876f40a47a939e4668db",
        "2220cd9936ebdf908d64c0571a4c8de83973eaa89c6778a64afec07de7c5e701",
        "8a55700bdaf62a00c208c7286531fd56ca321571f73f7620505a823ef5d4d0f1",
    }
)
_PRODUCT_ID: Final = re.compile(RELATION_GRAMMAR.product_id_pattern)
_MERGED_CELL: Final = re.compile(rb"<mergeCell\s")


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _merged_counts(path: Path, sheet_count: int) -> tuple[int, ...]:
    with zipfile.ZipFile(path) as archive:
        return tuple(
            len(_MERGED_CELL.findall(archive.read(f"xl/worksheets/sheet{index}.xml")))
            for index in range(1, sheet_count + 1)
        )


def _external_link_count(path: Path) -> int:
    with zipfile.ZipFile(path) as archive:
        return sum(
            name.startswith("xl/externalLinks/externalLink")
            and name.endswith(".xml")
            and "/_rels/" not in name
            for name in archive.namelist()
        )


def _cell_at(worksheet: Worksheet, coordinate: str) -> Cell | MergedCell:
    row, column = coordinate_to_tuple(coordinate)
    return worksheet.cell(row=row, column=column)


def _cell_text(cell: Cell | MergedCell) -> str:
    if cell.data_type == "f":
        message = f"formula relation cell: {cell.coordinate}"
        raise FixtureError(message)
    value = cell.value
    return "" if value is None else str(value).strip()


def _relation_counts(workbook_path: Path) -> tuple[int, int]:
    with closing(
        load_workbook(workbook_path, read_only=True, data_only=False, keep_links=True)
    ) as workbook:
        worksheet = next(
            sheet
            for sheet in workbook.worksheets
            if sheet.title == RELATION_GRAMMAR.sheet
        )
        for expected in RELATION_GRAMMAR.headers:
            if _cell_text(_cell_at(worksheet, expected.coordinate)) != expected.value:
                message = f"relationship header drift: {expected.coordinate}"
                raise FixtureError(message)
        parent = _cell_text(_cell_at(worksheet, RELATION_GRAMMAR.parent.coordinate))
        children = tuple(
            _cell_text(_cell_at(worksheet, coordinate))
            for coordinate in RELATION_GRAMMAR.child_cells
        )
        unbound = tuple(
            _cell_text(_cell_at(worksheet, coordinate))
            for coordinate in RELATION_GRAMMAR.unbound_option_cells
        )
    if parent != RELATION_GRAMMAR.parent.value:
        message = f"relationship parent drift: {RELATION_GRAMMAR.parent.coordinate}"
        raise FixtureError(message)
    if not _PRODUCT_ID.fullmatch(parent):
        message = "invalid relationship parent ID"
        raise FixtureError(message)
    if any(not _PRODUCT_ID.fullmatch(value) for value in (*children, *unbound)):
        message = "invalid relationship child ID"
        raise FixtureError(message)
    if len(children) != RELATION_GRAMMAR.curated_relationship_count:
        message = "workbook relationship count changed"
        raise FixtureError(message)
    if len(unbound) != RELATION_GRAMMAR.unbound_option_count:
        message = "workbook unbound option count changed"
        raise FixtureError(message)
    if len(set(children)) != len(children) or parent in children:
        message = "duplicate or self-linked relationship"
        raise FixtureError(message)
    return len(children), len(unbound)


def _extract_workbook(path: Path) -> WorkbookFacts:
    sha256 = _sha256(path)
    if sha256 not in EXPECTED_SHA256:
        message = f"source workbook SHA changed: {path.name}"
        raise FixtureError(message)
    with closing(
        load_workbook(path, read_only=True, data_only=False, keep_links=True)
    ) as workbook:
        merged_counts = _merged_counts(path, len(workbook.worksheets))
        sheets = tuple(
            SheetFacts(
                name=worksheet.title,
                dimension=worksheet.calculate_dimension(),
                rows=worksheet.max_row,
                columns=worksheet.max_column,
                formula_cells=sum(
                    cell.data_type == "f"
                    for row in worksheet.iter_rows()
                    for cell in row
                ),
                merged_ranges=merged_counts[index],
            )
            for index, worksheet in enumerate(workbook.worksheets)
        )
    curated, unbound = (
        _relation_counts(path) if sha256 == RELATION_GRAMMAR.source_sha256 else (0, 0)
    )
    totals = WorkbookTotals(
        formula_cells=sum(sheet.formula_cells for sheet in sheets),
        merged_ranges=sum(sheet.merged_ranges for sheet in sheets),
        external_links=_external_link_count(path),
        curated_relationships=curated,
        unbound_options=unbound,
    )
    return WorkbookFacts(
        filename=path.name,
        sha256=sha256,
        sheets=sheets,
        totals=totals,
    )


def _source_paths(source_dir: Path) -> tuple[Path, ...]:
    paths = tuple(
        sorted(source_dir.glob("*.xlsx"), key=lambda path: path.name.encode())
    )
    if len(paths) != len(EXPECTED_SHA256):
        message = f"expected exactly three XLSX workbooks, found {len(paths)}"
        raise FixtureError(message)
    return paths


def _manifest(paths: tuple[Path, ...]) -> Manifest:
    workbooks = tuple(_extract_workbook(path) for path in paths)
    if frozenset(item.sha256 for item in workbooks) != EXPECTED_SHA256:
        message = "source workbook SHA set changed"
        raise FixtureError(message)
    return Manifest(
        schema_version="workbook-manifest-v1",
        relation_grammar=RELATION_GRAMMAR,
        workbooks=workbooks,
    )


def _smoke_fixtures(paths: tuple[Path, ...]) -> tuple[SmokeFixture, SmokeFixture]:
    normalization: list[str] = []
    ranking: list[str] = []
    for path in paths:
        sha256 = _sha256(path)
        with closing(
            load_workbook(path, read_only=True, data_only=False, keep_links=True)
        ) as workbook:
            if sha256 == RELATION_GRAMMAR.source_sha256:
                worksheet = next(
                    sheet
                    for sheet in workbook.worksheets
                    if sheet.title == RELATION_GRAMMAR.sheet
                )
                for row in (9, 11, 12, 14, 21, 22, 28):
                    raw = _cell_text(worksheet.cell(row=row, column=4))
                    normalization.append(
                        f"{sha256}:{RELATION_GRAMMAR.sheet}:D{row}:{raw}"
                    )
            for worksheet in workbook.worksheets:
                if worksheet.title not in ("단가조사", "조달물품"):
                    continue
                for row in worksheet.iter_rows():
                    ranking.extend(
                        f"{sha256}:{worksheet.title}:{cell.coordinate}:{cell.value}"
                        for cell in row
                        if cell.data_type != "f"
                        and _PRODUCT_ID.fullmatch(str(cell.value))
                    )
    return (
        SmokeFixture(
            schema_version="normalization-workbook-smoke-v1",
            cases=tuple(normalization),
        ),
        SmokeFixture(
            schema_version="ranking-workbook-smoke-v1",
            cases=tuple(sorted(set(ranking), key=str.encode)),
        ),
    )


def build_fixture_bundle(source_dir: Path) -> FixtureBundle:
    """Extract every deterministic fixture from the pinned source set."""
    paths = _source_paths(source_dir)
    normalization, ranking = _smoke_fixtures(paths)
    return FixtureBundle(
        manifest=_manifest(paths),
        normalization=normalization,
        ranking=ranking,
    )


def verify_manifest(actual: Manifest, expected_path: Path) -> None:
    """Fail when any pinned workbook fact or relation grammar drifts."""
    expected = Manifest.model_validate_json(expected_path.read_text(encoding="utf-8"))
    for actual_item, expected_item in zip(
        actual.workbooks, expected.workbooks, strict=True
    ):
        if actual_item.sha256 != expected_item.sha256:
            message = "source workbook SHA changed"
            raise FixtureError(message)
        for actual_count, expected_count, message in (
            (
                actual_item.totals.formula_cells,
                expected_item.totals.formula_cells,
                "workbook formula count changed",
            ),
            (
                actual_item.totals.external_links,
                expected_item.totals.external_links,
                "workbook external link count changed",
            ),
            (
                actual_item.totals.curated_relationships,
                expected_item.totals.curated_relationships,
                "workbook relationship count changed",
            ),
        ):
            if actual_count != expected_count:
                raise FixtureError(message)
    if actual != expected:
        message = "workbook manifest changed"
        raise FixtureError(message)
